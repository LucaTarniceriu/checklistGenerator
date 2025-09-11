from openpyxl import load_workbook
from main.utils.DoorComponent import DoorComponent
import os
from .Site import Site

class Door:
    def __init__(self, site = "empty", productType = "empty", produs = "empty", anFabricatie = 0, nr = "empty", dimensiuni = "empty", tip = "empty", titluTabel = "empty", nrCanate = "", model="", data_inspecite="", tehnician="", lipsuri=False, informare=False):
        self.site = site

        self.productType = productType
        self.produs = produs
        self.anFabricatie = anFabricatie
        self.nr = nr
        self.dimensiuni = dimensiuni
        self.tip = tip
        self.fileName=""
        self.nrComponente=0

        self.titluTabel = titluTabel

        self.nrCanate = nrCanate
        self.model = model

        self.id = int
        self.componente = []
        self.data_inspectiei = data_inspecite
        self.tehnician = tehnician

        self.lipsuri = lipsuri
        self.informare = informare


    def setFileName(self):
        # a se apela dupa setarea atributului "productType"
        if self.productType == "1": #Antifoc
            self.fileName = "Antifoc.xlsx"
            self.nrComponente = 26
        elif self.productType == "2": #Automata
            self.fileName = "Automata.xlsx"
            self.nrComponente = 22
        elif self.productType == "3": #Burduf
            self.fileName = "Burduf.xlsx"
            self.nrComponente = 16
        elif self.productType == "4": #Metalica
            self.fileName = "Metalica.xlsx"
            self.nrComponente = 14
        elif self.productType == "5": #Rampa
            self.fileName = "Rampa.xlsx"
            self.nrComponente = 16
        elif self.productType == "6": #Rapida
            self.fileName = "Rapida.xlsx"
            self.nrComponente = 17
        elif self.productType == "7": #Sectionala
            self.fileName = "Sectionala.xlsx"
            self.nrComponente = 20
        else:
            print("eroare selectare usa")


    def setComponents(self):
        # a se apela dupa apelarea "setFileName"
        numeComponente = []
        wb = load_workbook(filename=os.path.join(Site.input_dir, self.fileName))
        ws = wb.active
        for row in range(14, 14 + self.nrComponente):
            nume = ws['C' + str(row)].value
            numeComponente.append(ws['C' + str(row)].value )
            # se extrage numele componentelor de verificat (a doua coloana a tabelului)
            self.componente.append(DoorComponent(nume, row-13))


    def __repr__(self):
        return self.productType + " " + self.produs + " " + str(self.anFabricatie) + " " + str(self.site)
