import subprocess, os, webbrowser, shutil
from openpyxl import load_workbook
class Site:

    #!! change this for Windows
    office_path = "/usr/bin/soffice"
    output_dir = "/home/themartianx/Documents/checklists/" #this folder has to be empty
    final_output_dir = "/home/themartianx/Documents/checklistsFinal/" #this folder has to be empty
    input_dir = "/home/themartianx/PycharmProjects/checklistGenerator/checklistGenerator/xlsx/"

    def __init__(self, contract, beneficiar, locatie, nrComanda):
        self.contract = contract
        self.beneficiar = beneficiar
        self.locatie = locatie
        self.nrComanda = nrComanda
        self.doors = []

    def addDoor(self, door):
        self.doors.append(door)

    def __repr__(self):
        return self.beneficiar+" "+self.locatie

    @staticmethod
    def fillFile(siteObject, doorObject, doorCount, reprezentant=" ", boolLipsuri="□", boolInformat="□"):
        #!! Modify the directory for Windows
        copyDir = "/home/themartianx/checklistTemp"
        newName = doorObject.oras+siteObject.locatie+"("+str(doorCount)+")"+".xlsx"
        copiedFile = os.path.join(copyDir, newName)
        shutil.copy(Site.input_dir+doorObject.fileName, copiedFile)
        wb = load_workbook(filename=copiedFile)
        ws = wb.active
        ws['E1'] = siteObject.contract
        ws['D2'] = doorObject.produs
        ws['D3'] = siteObject.beneficiar
        ws['D4'] = siteObject.locatie
        ws['D5'] = doorObject.anFabricatie
        ws['D6'] = siteObject.nrComanda
        ws['D7'] = doorObject.nr
        ws['D8'] = doorObject.dimensiuni
        ws['D9'] = doorObject.tip

        ws['B12'] = doorObject.titluTabel
        ws['B47'] = boolLipsuri
        ws['B48'] = boolInformat
        ws['D49'] = doorObject.data_inspectiei
        ws['D50'] = doorObject.tehnician
        ws['D51'] = reprezentant


        for row in range(1, doorObject.nrComponente+1):
            for components in doorObject.componente:
                if components.nrcrt == row:
                    if components.verified:
                        ws['D'+ str(row+13)] = "✓"
                    else:
                        ws['D' + str(row + 13)] = " "

                    if components.broken:
                        ws['E'+ str(row+13)] = "✓"
                    else:
                        ws['E' + str(row + 13)] = " "

                    ws['F' + str(row + 13)] = components.number
                    ws['G' + str(row + 13)] = components.notes


        wb.save(copiedFile)
        Site.pdfExport(copiedFile)
        os.remove(copiedFile)

    @staticmethod
    def pdfExport(fileName):
        subprocess.run([
            Site.office_path,
            "--headless",
            "--convert-to",
            "pdf:calc_pdf_Export:{\"Magnification\":{\"type\":\"long\",\"value\":\"4\"}, \"Zoom\":{\"type\":\"long\",\"value\":\"65\"}, \"PageRange\":{\"type\":\"string\",\"value\":\"1\"}}",
            "--outdir", Site.output_dir,

            fileName
        ])
        # webbrowser.open_new( os.path.join(Site.output_dir, os.path.splitext(os.path.basename(fileName))[0] + ".pdf"))